from email.policy import default
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout, update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.password_validation import validate_password
from django.core.exceptions import ValidationError
from django.contrib import messages
from django.http import JsonResponse
from django.template import TemplateDoesNotExist
from django.views.decorators.csrf import csrf_exempt
from django.utils import timezone
from django.db.models import Q
import json
import secrets
from datetime import datetime

from .models import *
from django.core.paginator import Paginator
from django.db.models import Q

# ============== AUTHENTICATION VIEWS ==============

def manager_register(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        confirm_password = request.POST.get('confirm_password')
        phone = request.POST.get('phone')
        
        if password != confirm_password:
            messages.error(request, 'Passwords do not match!')
            return redirect('manager_register')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('manager_register')
        
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=password,
            user_type='manager',
            phone=phone,
            is_first_login=False
        )
        messages.success(request, 'Manager account created successfully! Please login.')
        return redirect('login')
    
    return render(request, 'auth/manager_register.html')
def user_login(request):
    if request.user.is_authenticated:
        if request.user.user_type == 'manager':
            return redirect('manager_dashboard')
        else:
            return redirect('employee_dashboard')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # Check if this is employee's first login
            if user.user_type == 'employee' and user.is_first_login:
                messages.info(request, 'This is your first login. Please change your password to continue.')
                return redirect('change_password')
            
            # Redirect based on user type
            if user.user_type == 'manager':
                return redirect('manager_dashboard')
            else:
                return redirect('employee_dashboard')
        else:
            messages.error(request, 'Invalid username or password!')
    
    return render(request, 'auth/login.html')

def user_logout(request):
    logout(request)
    messages.success(request, 'You have been logged out successfully!')
    return redirect('login')
@login_required
def change_password(request):
    if request.method == 'POST':
        old_password = request.POST.get('old_password')
        new_password = request.POST.get('new_password')
        confirm_password = request.POST.get('confirm_password')
        
        user = request.user
        
        # Basic validation
        if not old_password or not new_password or not confirm_password:
            messages.error(request, 'All fields are required!')
            return redirect('change_password')
        
        if new_password != confirm_password:
            messages.error(request, 'New passwords do not match!')
            return redirect('change_password')
        
        # Check if new password is same as old password
        if old_password == new_password:
            messages.error(request, 'New password cannot be same as old password!')
            return redirect('change_password')
        
        # Check if old password is correct
        if not user.check_password(old_password):
            messages.error(request, 'Current password is incorrect!')
            return redirect('change_password')
        
        # Validate new password strength
        try:
            validate_password(new_password, user)
        except ValidationError as e:
            for error in e.messages:
                messages.error(request, error)
            return redirect('change_password')
        
        # All validation passed, change password
        user.set_password(new_password)
        user.is_first_login = False
        user.save()
        update_session_auth_hash(request, user)  # Keep user logged in
        
        messages.success(request, 'Password changed successfully! You can now access all features.')
        
        # Redirect based on user type
        if user.user_type == 'manager':
            return redirect('manager_dashboard')
        else:
            return redirect('employee_dashboard')
    
    # Check if this is first login for styling
    is_first_login = request.user.is_authenticated and request.user.user_type == 'employee' and request.user.is_first_login
    
    context = {
        'is_first_login': is_first_login,
    }
    return render(request, 'auth/change_password.html', context)
# ============== MANAGER VIEWS ==============
from django.utils import timezone
from datetime import timedelta
from django.db.models import Count, Q
from django.core.paginator import Paginator
from django.contrib.auth.decorators import login_required
from django.shortcuts import render, redirect
from .models import CustomUser, Project, Hardware, HardwareAssignment, HardwareSerialEntry, HardwareType

@login_required
def manager_dashboard(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    # Get base data
    employees = CustomUser.objects.filter(user_type='employee', manager=request.user)
    projects = Project.objects.filter(created_by=request.user)
    hardware_assignments = HardwareAssignment.objects.filter(assigned_by=request.user)
    hardware_items = Hardware.objects.filter(created_by=request.user)
    
    # Hardware status counts
    available_count = hardware_items.filter(status='available').count()
    assigned_count = hardware_items.filter(status='assigned').count()
    in_use_count = hardware_items.filter(status='in_use').count()
    maintenance_count = hardware_items.filter(status='maintenance').count()
    total_hardware = hardware_items.count()
    
    # Find maximum count for chart scaling
    max_count = max(available_count, assigned_count, in_use_count, maintenance_count) if total_hardware > 0 else 100
    # Add 10% padding for better visualization
    chart_max = int(max_count * 1.1) + 5
    
    # Calculate scale values for Y-axis
    scale_values = [
        chart_max,
        int(chart_max * 0.75),
        int(chart_max * 0.5),
        int(chart_max * 0.25),
        0
    ]
    
    # Calculate percentages and heights for chart
    hardware_status = [
        {
            'name': 'Available',
            'count': available_count,
            'percentage': (available_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((available_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'success'
        },
        {
            'name': 'Assigned',
            'count': assigned_count,
            'percentage': (assigned_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((assigned_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'warning'
        },
        {
            'name': 'In Use',
            'count': in_use_count,
            'percentage': (in_use_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((in_use_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'info'
        },
        {
            'name': 'Maintenance',
            'count': maintenance_count,
            'percentage': (maintenance_count / total_hardware * 100) if total_hardware > 0 else 0,
            'height': int((maintenance_count / max_count) * 100) if max_count > 0 else 0,
            'color': 'danger'
        },
    ]
    
    # Recent Activities
    recent_activities = []
    
    # Get recent assignments (2 items)
    recent_assignments = hardware_assignments.order_by('-assigned_date')[:2]
    for assignment in recent_assignments:
        recent_activities.append({
            'icon': 'clipboard-check',
            'color': 'primary',
            'title': 'New assignment created',
            'description': f'{assignment.employee.get_full_name() or assignment.employee.username} - {assignment.project.project_name}',
            'timestamp': assignment.assigned_date,
            'type': 'assignment'
        })
    
    # Get recent employee registrations (1 item)
    recent_employees = employees.order_by('-date_joined')[:1]
    for employee in recent_employees:
        recent_activities.append({
            'icon': 'person-plus',
            'color': 'success',
            'title': 'New employee registered',
            'description': f'{employee.get_full_name() or employee.username} joined',
            'timestamp': employee.date_joined,
            'type': 'employee'
        })
    
    # Get recent hardware verifications (1 item)
    try:
        recent_verifications = HardwareSerialEntry.objects.filter(
            assignment_item__assignment__assigned_by=request.user,
            verified=True
        ).order_by('-verified_at')[:1]
        
        for verification in recent_verifications:
            recent_activities.append({
                'icon': 'shield-check',
                'color': 'info',
                'title': 'Hardware verified',
                'description': f'{verification.assignment_item.hardware.hardware_type.name} verified',
                'timestamp': verification.verified_at,
                'type': 'verification'
            })
    except:
        pass
    
    # Sort by timestamp (most recent first) and limit to 4 items
    recent_activities.sort(key=lambda x: x['timestamp'], reverse=True)
    recent_activities = recent_activities[:4]
    
    # Calculate growth percentages
    thirty_days_ago = timezone.now() - timedelta(days=30)
    
    last_month_employees = employees.filter(date_joined__lt=thirty_days_ago).count()
    if last_month_employees > 0:
        employee_growth = ((employees.count() - last_month_employees) / last_month_employees * 100)
    else:
        employee_growth = 12  # Default value
    
    last_month_projects = projects.filter(created_at__lt=thirty_days_ago).count()
    if last_month_projects > 0:
        project_growth = ((projects.count() - last_month_projects) / last_month_projects * 100)
    else:
        project_growth = 8  # Default value
    
    last_month_assignments = hardware_assignments.filter(assigned_date__lt=thirty_days_ago).count()
    if last_month_assignments > 0:
        assignment_growth = ((hardware_assignments.count() - last_month_assignments) / last_month_assignments * 100)
    else:
        assignment_growth = 15  # Default value
    
    last_month_hardware = hardware_items.filter(created_at__lt=thirty_days_ago).count()
    if last_month_hardware > 0:
        hardware_growth = ((hardware_items.count() - last_month_hardware) / last_month_hardware * 100)
    else:
        hardware_growth = 5  # Default value
    
    context = {
        'total_employees': employees.count(),
        'total_projects': projects.count(),
        'active_assignments': hardware_assignments.filter(actual_return_date__isnull=True).count(),
        'hardware_count': total_hardware,
        'employees': employees[:5],
        'projects': projects[:5],
        'assignments': hardware_assignments.order_by('-assigned_date')[:5],
        'hardware_status': hardware_status,
        'available_count': available_count,
        'assigned_count': assigned_count,
        'in_use_count': in_use_count,
        'maintenance_count': maintenance_count,
        'recent_activities': recent_activities,
        'employee_growth': round(employee_growth, 1),
        'project_growth': round(project_growth, 1),
        'assignment_growth': round(assignment_growth, 1),
        'hardware_growth': round(hardware_growth, 1),
        'chart_max': chart_max,
        'scale_values': scale_values,
        'max_count': max_count,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/dashboard.html', context)

from django.core.mail import send_mail
from django.conf import settings

@login_required
def create_employee(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        username = request.POST.get('username')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        
        # Validate name
        if not name or len(name.strip()) < 2:
            messages.error(request, 'Please enter a valid name!')
            return redirect('create_employee')
        
        if CustomUser.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
            return redirect('create_employee')
        
        if CustomUser.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
            return redirect('create_employee')
        
        # Generate a temporary password
        import secrets
        import string
        alphabet = string.ascii_letters + string.digits
        temp_password = ''.join(secrets.choice(alphabet) for i in range(10))
        
        # Split name into first and last name
        name_parts = name.strip().split(' ', 1)
        first_name = name_parts[0]
        last_name = name_parts[1] if len(name_parts) > 1 else ''
        
        # Create employee user
        user = CustomUser.objects.create_user(
            username=username,
            email=email,
            password=temp_password,
            user_type='employee',
            manager=request.user,
            phone=phone,
            is_first_login=True,
            first_name=first_name,
            last_name=last_name
        )
        
        # Send email with credentials
        try:
            send_mail(
                subject='Your Eduquity Hardware Management Account Credentials',
                message=f'''
Dear {name},

Welcome to Eduquity Hardware Management System!

Your account has been created successfully. Here are your login credentials:

Username: {username}
Temporary Password: {temp_password}
Login URL: http://127.0.0.1:8000/login/

Important Instructions:
1. This is your first login - you MUST change your password immediately
2. Use the temporary password above to login
3. After login, you will be prompted to change your password
4. Keep your credentials secure

For security reasons, please change your password immediately after first login.

Best regards,
Eduquity Hardware Management Team
                ''',
                html_message=f'''
<!DOCTYPE html>
<html>
<head>
    <style>
        body {{ font-family: Arial, sans-serif; line-height: 1.6; color: #333; }}
        .container {{ max-width: 600px; margin: 0 auto; padding: 20px; }}
        .header {{ background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; padding: 20px; text-align: center; border-radius: 5px 5px 0 0; }}
        .content {{ background: #f9f9f9; padding: 30px; border: 1px solid #ddd; border-top: none; border-radius: 0 0 5px 5px; }}
        .credentials {{ background: #e8f4fc; border: 2px solid #3498db; padding: 15px; margin: 20px 0; border-radius: 5px; }}
        .warning {{ background: #fff3cd; border: 1px solid #ffc107; padding: 15px; margin: 20px 0; border-radius: 5px; color: #856404; }}
        .footer {{ margin-top: 30px; padding-top: 20px; border-top: 1px solid #ddd; color: #666; font-size: 12px; }}
        .btn {{ display: inline-block; padding: 10px 20px; background: linear-gradient(90deg, #2c3e50 0%, #3498db 100%); color: white; text-decoration: none; border-radius: 5px; margin: 10px 0; }}
    </style>
</head>
<body>
    <div class="container">
        <div class="header">
            <h2>Welcome to Eduquity Hardware Management</h2>
        </div>
        <div class="content">
            <p>Dear <strong>{name}</strong>,</p>
            
            <p>Your account has been created successfully in the Eduquity Hardware Management System.</p>
            
            <div class="credentials">
                <h3>Your Login Credentials:</h3>
                <p><strong>Full Name:</strong> {name}</p>
                <p><strong>Username:</strong> {username}</p>
                <p><strong>Temporary Password:</strong> <code style="background: #fff; padding: 5px 10px; border-radius: 3px; font-size: 14px;">{temp_password}</code></p>
                <p><strong>Login URL:</strong> <a href="http://127.0.0.1:8000/login/">http://127.0.0.1:8000/login/</a></p>
                <a href="http://127.0.0.1:8000/login/" class="btn">Login Now</a>
            </div>
            
            <div class="warning">
                <h4>⚠️ Important Security Instructions:</h4>
                <ol>
                    <li>This is your <strong>first login</strong> - you MUST change your password immediately</li>
                    <li>Use the temporary password above to login</li>
                    <li>After login, you will be prompted to change your password</li>
                    <li>Keep your credentials secure and do not share with anyone</li>
                </ol>
            </div>
            
            <p><strong>About the System:</strong><br>
            The Eduquity Hardware Management System allows you to:
            <ul>
                <li>View your hardware assignments</li>
                <li>Enter serial numbers of assigned hardware</li>
                <li>Track hardware status</li>
                <li>Communicate with your manager</li>
            </ul>
            </p>
            
            <div class="footer">
                <p><strong>Eduquity Hardware Management Team</strong><br>
                Established in 2000 - Thought-leader in the Indian assessment industry</p>
                <p><em>This is an automated email. Please do not reply to this message.</em></p>
            </div>
        </div>
    </div>
</body>
</html>
                ''',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[email],
                fail_silently=False,
            )
            
            messages.success(request, f'Employee account created successfully for {name}! Login credentials have been sent to {email}.')
            
            # Also show the password to manager (for emergency cases)
            messages.info(request, f'Temporary password for {name}: {temp_password}')
            
        except Exception as e:
            # If email fails, still create user but show warning
            messages.warning(request, f'Employee account created for {name} but email could not be sent. Error: {str(e)}')
            messages.info(request, f'Temporary password for {name}: {temp_password}')
        
        return redirect('create_employee')
    
    # Get ALL employees for this manager (for statistics)
    all_employees = CustomUser.objects.filter(
        user_type='employee', 
        manager=request.user
    ).order_by('-date_joined')
    
    # Get recently created employees (last 5 for display)
    recent_employees = all_employees[:5]
    
    # Calculate statistics from ALL employees
    total_employees = all_employees.count()
    active_employees = all_employees.filter(is_first_login=False).count()
    pending_employees = all_employees.filter(is_first_login=True).count()
    
    context = {
        'recent_employees': recent_employees,
        'total_employees': total_employees,
        'active_employees': active_employees,
        'pending_employees': pending_employees,
    }
    return render(request, 'manager/create_employee.html', context)

@login_required
def delete_employee(request, employee_id):
    """Delete an employee account"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employee = get_object_or_404(
        CustomUser, 
        id=employee_id, 
        user_type='employee', 
        manager=request.user
    )
    
    if request.method == 'POST':
        employee_name = employee.get_full_name() or employee.username
        
        # Check if employee has any active assignments
        active_assignments = HardwareAssignment.objects.filter(
            employee=employee,
            actual_return_date__isnull=True
        ).exists()
        
        if active_assignments:
            messages.error(
                request, 
                f'Cannot delete {employee_name} because they have active hardware assignments. Please return all hardware first.'
            )
            return redirect('create_employee')
        
        # Delete the employee
        employee.delete()
        messages.success(request, f'Employee {employee_name} has been deleted successfully.')
        return redirect('create_employee')
    
    return redirect('create_employee')


@login_required
def employee_list(request):
    """View all employees with management options"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employees = CustomUser.objects.filter(
        user_type='employee', 
        manager=request.user
    ).order_by('-date_joined')
    
    # Get assignment counts for each employee
    for emp in employees:
        emp.active_assignments = HardwareAssignment.objects.filter(
            employee=emp,
            actual_return_date__isnull=True
        ).count()
        emp.total_assignments = HardwareAssignment.objects.filter(
            employee=emp
        ).count()
    
    context = {
        'employees': employees,
        'total_employees': employees.count(),
        'active_employees': employees.filter(is_first_login=False).count(),
        'pending_employees': employees.filter(is_first_login=True).count(),
    }
    return render(request, 'manager/employee_list.html', context)    

@login_required
def create_project(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    if request.method == 'POST':
        project_id = request.POST.get('project_id')
        project_name = request.POST.get('project_name')
        description = request.POST.get('description')
        start_date = request.POST.get('start_date')
        end_date = request.POST.get('end_date')
        location = request.POST.get('location')
        
        if Project.objects.filter(project_id=project_id).exists():
            messages.error(request, 'Project ID already exists!')
            return redirect('create_project')
        
        project = Project.objects.create(
            project_id=project_id,
            project_name=project_name,
            description=description,
            start_date=start_date,
            end_date=end_date,
            location=location,
            created_by=request.user
        )
        
        messages.success(request, 'Project created successfully!')
        return redirect('create_project')
    
    return render(request, 'manager/create_project.html')

# ============== HARDWARE MANAGEMENT VIEWS ==============
from django.core.paginator import Paginator
from django.db.models import Q
from django.core.paginator import Paginator
from django.db.models import Q, Count

@login_required
def manage_hardware(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    hardware_items = Hardware.objects.filter(created_by=request.user)
    
    # Get count of items for each hardware type
    hardware_items_count = hardware_items.values('hardware_type').annotate(
        count=Count('id')
    ).order_by()
    
    # Convert to dictionary for easy lookup in template
    hardware_items_count_dict = {}
    for item in hardware_items_count:
        hardware_items_count_dict[item['hardware_type']] = item['count']
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        hardware_items = hardware_items.filter(
            Q(serial_number__icontains=search_query) |
            Q(model_name__icontains=search_query) |
            Q(brand__icontains=search_query) |
            Q(hardware_type__name__icontains=search_query)
        )
    
    # Filter by hardware type
    type_filter = request.GET.get('type', '')
    if type_filter and type_filter != 'all':
        hardware_items = hardware_items.filter(hardware_type_id=type_filter)
    
    # Filter by status
    status_filter = request.GET.get('status', '')
    if status_filter and status_filter != 'all':
        hardware_items = hardware_items.filter(status=status_filter)
    
    # Get statistics
    available_count = hardware_items.filter(status='available').count()
    assigned_count = hardware_items.filter(status='assigned').count()
    in_use_count = hardware_items.filter(status='in_use').count()
    maintenance_count = hardware_items.filter(status='maintenance').count()
    
    # Pagination
    paginator = Paginator(hardware_items, 10)
    page_number = request.GET.get('page', 1)
    page_obj = paginator.get_page(page_number)
    
    # Create a custom template filter or pass the dictionary directly
    from django.template.defaulttags import register
    register.filter('get_item', lambda d, key: d.get(key, 0))
    
    context = {
        'hardware_types': hardware_types,
        'hardware_items': page_obj,
        'hardware_items_count_dict': hardware_items_count_dict,
        'available_count': available_count,
        'assigned_count': assigned_count,
        'in_use_count': in_use_count,
        'maintenance_count': maintenance_count,
        'search_query': search_query,
        'type_filter': type_filter,
        'status_filter': status_filter,
        'paginator': paginator,
        'page_obj': page_obj,
    }
    return render(request, 'manager/manage_hardware.html', context)

@login_required
def add_hardware(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        hardware_type_id = request.POST.get('hardware_type')
        serial_number = request.POST.get('serial_number')
        model_name = request.POST.get('model_name')
        brand = request.POST.get('brand')
        specifications = request.POST.get('specifications')
        purchase_date = request.POST.get('purchase_date')
        
        # Check if serial number already exists
        if Hardware.objects.filter(serial_number=serial_number).exists():
            messages.error(request, 'Serial number already exists!')
            return redirect('add_hardware')
        
        try:
            hardware_type = HardwareType.objects.get(id=hardware_type_id)
        except HardwareType.DoesNotExist:
            messages.error(request, 'Invalid hardware type selected!')
            return redirect('add_hardware')
        
        hardware = Hardware.objects.create(
            hardware_type=hardware_type,
            serial_number=serial_number,
            model_name=model_name,
            brand=brand,
            specifications=specifications,
            purchase_date=purchase_date if purchase_date else None,
            status='available',
            created_by=request.user
        )
        
        messages.success(request, f'{hardware_type.name} added successfully!')
        return redirect('manage_hardware')
    
    context = {'hardware_types': hardware_types}
    return render(request, 'manager/add_hardware.html', context)

@login_required
def edit_hardware(request, hardware_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware = get_object_or_404(Hardware, id=hardware_id, created_by=request.user)
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        hardware_type_id = request.POST.get('hardware_type')
        serial_number = request.POST.get('serial_number')
        model_name = request.POST.get('model_name')
        brand = request.POST.get('brand')
        specifications = request.POST.get('specifications')
        purchase_date = request.POST.get('purchase_date')
        status = request.POST.get('status')
        
        # Check if serial number already exists (excluding current hardware)
        if Hardware.objects.filter(serial_number=serial_number).exclude(id=hardware_id).exists():
            messages.error(request, 'Serial number already exists!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        try:
            hardware.hardware_type = HardwareType.objects.get(id=hardware_type_id)
        except HardwareType.DoesNotExist:
            messages.error(request, 'Invalid hardware type selected!')
            return redirect('edit_hardware', hardware_id=hardware_id)
        
        hardware.serial_number = serial_number
        hardware.model_name = model_name
        hardware.brand = brand
        hardware.specifications = specifications
        hardware.purchase_date = purchase_date if purchase_date else None
        hardware.status = status
        hardware.save()
        
        messages.success(request, 'Hardware updated successfully!')
        return redirect('manage_hardware')
    
    context = {
        'hardware': hardware,
        'hardware_types': hardware_types,
    }
    return render(request, 'manager/edit_hardware.html', context)

@login_required
def delete_hardware(request, hardware_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware = get_object_or_404(Hardware, id=hardware_id, created_by=request.user)
    
    if request.method == 'POST':
        # Check if hardware is assigned
        if hardware.status == 'assigned' or hardware.status == 'in_use':
            messages.error(request, 'Cannot delete hardware that is currently assigned!')
            return redirect('manage_hardware')
        
        hardware.delete()
        messages.success(request, 'Hardware deleted successfully!')
        return redirect('manage_hardware')
    
    context = {'hardware': hardware}
    return render(request, 'manager/delete_hardware.html', context)

@login_required
def manage_hardware_types(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    hardware_types = HardwareType.objects.all()
    
    if request.method == 'POST':
        name = request.POST.get('name')
        description = request.POST.get('description')
        
        if HardwareType.objects.filter(name=name).exists():
            messages.error(request, 'Hardware type already exists!')
        else:
            HardwareType.objects.create(
                name=name,
                description=description
            )
            messages.success(request, 'Hardware type added successfully!')
        
        return redirect('manage_hardware_types')
    
    context = {'hardware_types': hardware_types}
    return render(request, 'manager/manage_hardware_types.html', context)

# ============== ASSIGNMENT VIEWS ==============

@login_required
def create_assignment(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    employees = CustomUser.objects.filter(user_type='employee', manager=request.user, is_active=True)
    projects = Project.objects.filter(created_by=request.user)
    hardware_types = HardwareType.objects.all()
    
    # Get available hardware items for verification
    for hw_type in hardware_types:
        hw_type.available_serials = list(hw_type.hardware_set.filter(
            status='available',
            created_by=request.user
        ).values_list('serial_number', flat=True))
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee')
        project_id = request.POST.get('project')
        expected_return_date = request.POST.get('expected_return_date')
        exam_city = request.POST.get('exam_city')
        notes = request.POST.get('notes')
        
        # Get hardware type IDs and serial numbers
        hardware_type_ids = request.POST.getlist('hardware_type[]')
        serial_numbers = request.POST.getlist('serial_numbers[]')
        
        # Validate required fields
        if not all([employee_id, project_id, expected_return_date, exam_city]):
            messages.error(request, 'Please fill in all required fields!')
            return redirect('create_assignment')
        
        # Validate hardware items
        if not hardware_type_ids or not serial_numbers:
            messages.error(request, 'Please add at least one hardware item!')
            return redirect('create_assignment')
        
        try:
            employee = CustomUser.objects.get(id=employee_id, manager=request.user)
            project = Project.objects.get(id=project_id, created_by=request.user)
        except (CustomUser.DoesNotExist, Project.DoesNotExist):
            messages.error(request, 'Invalid employee or project selected!')
            return redirect('create_assignment')
        
        # Verify and get hardware items
        hardware_items = []
        errors = []
        
        for i, (type_id, serial) in enumerate(zip(hardware_type_ids, serial_numbers)):
            serial = serial.strip()
            
            try:
                hardware = Hardware.objects.get(
                    hardware_type_id=type_id,
                    serial_number=serial,
                    status='available',
                    created_by=request.user
                )
                hardware_items.append(hardware)
            except Hardware.DoesNotExist:
                errors.append(f"Row {i+1}: Hardware '{serial}' not found or not available")
        
        if errors:
            messages.error(request, 'Serial number verification failed:\n' + '\n'.join(errors))
            return redirect('create_assignment')
        
        # Create assignment
        assignment = HardwareAssignment.objects.create(
            employee=employee,
            project=project,
            assigned_by=request.user,
            expected_return_date=expected_return_date,
            exam_city=exam_city,
            notes=notes
        )
        
        # Create assignment items and update hardware status
        for hardware in hardware_items:
            HardwareAssignmentItem.objects.create(
                assignment=assignment,
                hardware=hardware,
                quantity=1,
                condition_at_assignment='Assigned for exam duty'
            )
            hardware.status = 'assigned'
            hardware.save()
        
        messages.success(request, f'Assignment created successfully with {len(hardware_items)} hardware item(s)!')
        return redirect('view_assignments')
    
    context = {
        'employees': employees,
        'projects': projects,
        'hardware_types': hardware_types,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/create_assignment.html', context)

from django.db.models import Count, Q
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.utils import timezone

from django.utils import timezone
from datetime import timedelta
from django.db.models import Q, Count

@login_required
def view_assignments(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    # Base queryset - ALL assignments
    all_assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user
    ).order_by('-assigned_date')
    
    # Filter for pending returns if requested
    today = timezone.now().date()
    due_soon_date = today + timedelta(days=3)
    
    # This is the queryset that will be displayed in the table
    if request.GET.get('pending') == '1':
        assignments = all_assignments.filter(
            actual_return_date__isnull=True,
            expected_return_date__lte=due_soon_date
        )
    else:
        assignments = all_assignments  # Show all assignments
    
    # Calculate statistics from ALL assignments
    total_assignments = all_assignments.count()
    active_assignments = all_assignments.filter(actual_return_date__isnull=True).count()
    returned_assignments = all_assignments.filter(actual_return_date__isnull=False).count()
    
    # Calculate OVERDUE assignments (past expected return date)
    overdue_count = all_assignments.filter(
        actual_return_date__isnull=True,
        expected_return_date__lt=today  # Less than today, not including today
    ).count()
    
    # Calculate DUE SOON assignments (today and next 3 days)
    due_soon_count = all_assignments.filter(
        actual_return_date__isnull=True,
        expected_return_date__gte=today,  # Today or future
        expected_return_date__lte=due_soon_date
    ).count()
    
    # PENDING RETURN = Overdue + Due Soon
    pending_return_count = overdue_count + due_soon_count
    
    # Debug - Print to console
    print(f"=== DEBUG ===")
    print(f"Today: {today}")
    print(f"Due soon date: {due_soon_date}")
    print(f"Active assignments: {active_assignments}")
    print(f"Overdue count: {overdue_count}")
    print(f"Due soon count: {due_soon_count}")
    print(f"Pending return count: {pending_return_count}")
    print(f"=============")
    
    # For debugging - list all active assignments
    active_assignments_list = all_assignments.filter(actual_return_date__isnull=True)
    for assignment in active_assignments_list:
        status = "OVERDUE" if assignment.expected_return_date < today else "DUE SOON" if assignment.expected_return_date <= due_soon_date else "FUTURE"
        print(f"Assignment {assignment.id}: Expected: {assignment.expected_return_date}, Status: {status}")
    
    # Get unique exam cities count
    unique_cities = all_assignments.exclude(
        exam_city__isnull=True
    ).exclude(
        exam_city__exact=''
    ).values('exam_city').distinct().count()
    
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'active_assignments': active_assignments,
        'returned_assignments': returned_assignments,
        'pending_return_count': pending_return_count,
        'overdue_count': overdue_count,
        'due_soon_count': due_soon_count,
        'unique_cities': unique_cities,
        'today': today,
        'due_soon_date': due_soon_date,
    }
    return render(request, 'manager/view_assignments.html', context)

@login_required
def assignment_details(request, assignment_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment, 
        id=assignment_id, 
        assigned_by=request.user
    )
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    
    # Check serial entry status
    for item in items:
        item.has_serial_entry = hasattr(item, 'serial_entry')
        if item.has_serial_entry:
            item.serial_entry_status = item.serial_entry.verified
    
    context = {
        'assignment': assignment,
        'items': items,
    }
    return render(request, 'manager/assignment_details.html', context)
@login_required
def return_assignment(request, assignment_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, assigned_by=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    
    if request.method == 'POST':
        # Get verification data from form
        verification_status = []
        all_verified = True
        mismatch_count = 0
        missing_count = 0
        
        for item in items:
            returned_serial = request.POST.get(f'returned_serial_{item.id}', '').strip()
            condition_notes = request.POST.get(f'condition_notes_{item.id}', '')
            
            if not returned_serial:
                missing_count += 1
                all_verified = False
                verification_status.append({
                    'item': item,
                    'status': 'missing',
                    'message': 'Return serial number not entered'
                })
                continue
            
            # Check if returned serial matches assigned serial
            is_match = (returned_serial == item.hardware.serial_number)
            
            if not is_match:
                mismatch_count += 1
                all_verified = False
                verification_status.append({
                    'item': item,
                    'status': 'mismatch',
                    'message': f'Returned serial "{returned_serial}" does not match assigned serial "{item.hardware.serial_number}"'
                })
            
            # Update condition notes for the item
            item.condition_at_return = condition_notes
            item.save()
        
        # If verification fails, show error and return to form
        if not all_verified:
            error_messages = []
            if missing_count > 0:
                error_messages.append(f'{missing_count} item(s) missing return serial numbers')
            if mismatch_count > 0:
                error_messages.append(f'{mismatch_count} item(s) have mismatched serial numbers')
            
            messages.error(request, 'Return verification failed: ' + '; '.join(error_messages))
            
            # Store verification status in session for display
            request.session['verification_status'] = verification_status
            return redirect('return_assignment', assignment_id=assignment.id)
        
        # All verification passed - process return
        for item in items:
            # Update hardware status to available
            item.hardware.status = 'available'
            item.hardware.save()
        
        # Mark assignment as returned
        assignment.actual_return_date = timezone.now().date()
        assignment.save()
        
        messages.success(request, f'Assignment returned successfully! All {items.count()} hardware items verified and marked as available.')
        return redirect('view_assignments')
    
    # Check if there's verification status in session (for displaying errors)
    verification_status = request.session.pop('verification_status', [])
    
    context = {
        'assignment': assignment,
        'items': items,
        'verification_status': verification_status,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/return_assignment.html', context)

# ============== SERIAL NUMBER VIEWS ==============
@login_required
def view_serial_entries(request):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    # Get all assignments with serial entries
    assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).prefetch_related(
        'hardwareassignmentitem_set__hardware__hardware_type',
        'hardwareassignmentitem_set__serial_entry__entered_by',
        'hardwareassignmentitem_set__serial_entry__verified_by'
    ).order_by('-assigned_date')
    
    # Process verification stats for each assignment
    total_verified = 0
    total_matched = 0
    total_mismatch = 0
    total_pending = 0
    
    for assignment in assignments:
        items = assignment.hardwareassignmentitem_set.all()
        assignment.total_items = items.count()
        assignment.verified_count = 0
        assignment.matched_count = 0
        assignment.mismatch_count = 0
        assignment.pending_count = 0
        assignment.serial_entries = []
        
        for item in items:
            try:
                serial_entry = item.serial_entry
                # Check if serial numbers match
                is_match = serial_entry.serial_number == item.hardware.serial_number
                
                entry_data = {
                    'id': serial_entry.id,
                    'item_id': item.id,
                    'serial_number': serial_entry.serial_number,
                    'expected_serial': item.hardware.serial_number,
                    'hardware_type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'entered_by': serial_entry.entered_by,
                    'entered_at': serial_entry.entered_at,
                    'verified': serial_entry.verified,
                    'verified_by': serial_entry.verified_by,
                    'verified_at': serial_entry.verified_at,
                    'is_match': is_match,
                    'match_status': 'verified' if serial_entry.verified else ('matched' if is_match else 'mismatch'),
                }
                assignment.serial_entries.append(entry_data)
                
                if serial_entry.verified:
                    assignment.verified_count += 1
                    total_verified += 1
                else:
                    if is_match:
                        assignment.matched_count += 1
                        total_matched += 1
                    else:
                        assignment.mismatch_count += 1
                        total_mismatch += 1
                        
            except HardwareSerialEntry.DoesNotExist:
                # No serial entry yet
                entry_data = {
                    'id': None,
                    'item_id': item.id,
                    'serial_number': None,
                    'expected_serial': item.hardware.serial_number,
                    'hardware_type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'entered_by': None,
                    'entered_at': None,
                    'verified': False,
                    'verified_by': None,
                    'verified_at': None,
                    'is_match': False,
                    'match_status': 'pending',
                }
                assignment.serial_entries.append(entry_data)
                assignment.pending_count += 1
                total_pending += 1
    
    # Calculate totals
    total_assignments = assignments.count()
    
    context = {
        'assignments': assignments,
        'total_assignments': total_assignments,
        'total_verified': total_verified,
        'total_matched': total_matched,
        'total_mismatch': total_mismatch,
        'total_pending': total_pending,
    }
    return render(request, 'manager/view_serial_entries.html', context)


@login_required
def verify_serial_entry(request, entry_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    serial_entry = get_object_or_404(HardwareSerialEntry, id=entry_id)
    
    # Verify that this serial entry belongs to manager's assignment
    if serial_entry.assignment_item.assignment.assigned_by != request.user:
        messages.error(request, 'Unauthorized access!')
        return redirect('view_serial_entries')
    
    # Check if serial matches
    is_match = serial_entry.serial_number == serial_entry.assignment_item.hardware.serial_number
    
    if is_match:
        serial_entry.verified = True
        serial_entry.verified_by = request.user
        serial_entry.verified_at = timezone.now()
        serial_entry.save()
        
        # Update hardware status to "in_use"
        hardware = serial_entry.assignment_item.hardware
        hardware.status = 'in_use'
        hardware.save()
        
        messages.success(request, f'Serial entry verified successfully! Hardware is now marked as in use.')
    else:
        messages.error(request, 'Cannot verify - Serial number does not match the assigned hardware!')
    
    return redirect('view_serial_entries')


@login_required
def verify_all_employee_entries(request, assignment_id):
    """Single click verify all pending entries for an employee"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment,
        id=assignment_id,
        assigned_by=request.user,
        actual_return_date__isnull=True
    )
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    verified_count = 0
    skipped_count = 0
    mismatch_count = 0
    
    for item in items:
        try:
            serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
            # Check if serial matches
            is_match = serial_entry.serial_number == item.hardware.serial_number
            
            if not serial_entry.verified and is_match:
                serial_entry.verified = True
                serial_entry.verified_by = request.user
                serial_entry.verified_at = timezone.now()
                serial_entry.save()
                
                # Update hardware status
                item.hardware.status = 'in_use'
                item.hardware.save()
                verified_count += 1
            elif serial_entry.verified:
                skipped_count += 1
            elif not is_match:
                mismatch_count += 1
        except HardwareSerialEntry.DoesNotExist:
            continue
    
    employee_name = assignment.employee.get_full_name() or assignment.employee.username
    
    if verified_count > 0:
        messages.success(
            request, 
            f'Successfully verified {verified_count} hardware item(s) for {employee_name}!'
        )
        if mismatch_count > 0:
            messages.warning(
                request,
                f'Skipped {mismatch_count} item(s) with serial mismatch for {employee_name}.'
            )
        if skipped_count > 0:
            messages.info(
                request,
                f'{skipped_count} item(s) were already verified.'
            )
    else:
        if mismatch_count > 0:
            messages.error(
                request, 
                f'No items verified. Found {mismatch_count} item(s) with serial mismatch for {employee_name}.'
            )
        else:
            messages.warning(
                request, 
                f'No eligible items found for verification for {employee_name}.'
            )
    
    return redirect('view_serial_entries')
@login_required
def verify_serial_entry(request, entry_id):
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    serial_entry = get_object_or_404(HardwareSerialEntry, id=entry_id)
    
    # Verify that this serial entry belongs to manager's assignment
    if serial_entry.assignment_item.assignment.assigned_by != request.user:
        messages.error(request, 'Unauthorized access!')
        return redirect('view_serial_entries')
    
    # Check if serial matches
    is_match = serial_entry.serial_number == serial_entry.assignment_item.hardware.serial_number
    
    if is_match:
        serial_entry.verified = True
        serial_entry.verified_by = request.user
        serial_entry.verified_at = timezone.now()
        serial_entry.save()
        
        # Update hardware status to "in_use"
        hardware = serial_entry.assignment_item.hardware
        hardware.status = 'in_use'
        hardware.save()
        
        messages.success(request, f'Serial entry verified successfully! Hardware is now marked as in use.')
    else:
        messages.error(request, 'Cannot verify - Serial number does not match the assigned hardware!')
    
    return redirect('view_serial_entries')

@login_required
def verify_all_employee_entries(request, assignment_id):
    """Single click verify all pending entries for an employee"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment,
        id=assignment_id,
        assigned_by=request.user,
        actual_return_date__isnull=True
    )
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    verified_count = 0
    
    for item in items:
        try:
            serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
            # Only verify if serial matches and not already verified
            if not serial_entry.verified and serial_entry.serial_number == item.hardware.serial_number:
                serial_entry.verified = True
                serial_entry.verified_by = request.user
                serial_entry.verified_at = timezone.now()
                serial_entry.save()
                
                # Update hardware status
                item.hardware.status = 'in_use'
                item.hardware.save()
                verified_count += 1
        except HardwareSerialEntry.DoesNotExist:
            continue
    
    # FIXED: Use proper Python string formatting
    employee_name = assignment.employee.get_full_name() or assignment.employee.username
    
    if verified_count > 0:
        messages.success(
            request, 
            f'Successfully verified {verified_count} hardware item(s) for {employee_name}!'
        )
    else:
        messages.warning(
            request, 
            f'No eligible items found for verification for {employee_name}. Items must have matching serial numbers and not be already verified.'
        )
    
    return redirect('view_serial_entries')
@login_required
def manager_verification_status(request):
    """Manager dashboard to see verification status across all assignments"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignments = HardwareAssignment.objects.filter(
        assigned_by=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    total_verified = 0
    total_matched = 0
    total_mismatch = 0
    total_pending = 0
    
    # Add verification stats for each assignment
    for assignment in assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        total_items = items.count()
        verified_items = 0
        matched_items = 0
        mismatch_items = 0
        pending_items = 0
        
        for item in items:
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                if serial_entry.verified:
                    verified_items += 1
                else:
                    # Check if the entered serial matches the expected serial
                    if serial_entry.serial_number == item.hardware.serial_number:
                        matched_items += 1
                    else:
                        mismatch_items += 1
            except HardwareSerialEntry.DoesNotExist:
                pending_items += 1
        
        # Calculate percentages
        verified_percentage = (verified_items / total_items * 100) if total_items > 0 else 0
        matched_percentage = (matched_items / total_items * 100) if total_items > 0 else 0
        mismatch_percentage = (mismatch_items / total_items * 100) if total_items > 0 else 0
        pending_percentage = (pending_items / total_items * 100) if total_items > 0 else 0
        
        assignment.verification_stats = {
            'total': total_items,
            'verified': verified_items,
            'matched': matched_items,
            'mismatch': mismatch_items,
            'pending': pending_items,
            'verified_percentage': verified_percentage,
            'matched_percentage': matched_percentage,
            'mismatch_percentage': mismatch_percentage,
            'pending_percentage': pending_percentage,
            'progress': verified_percentage
        }
        
        # Update totals
        total_verified += verified_items
        total_matched += matched_items
        total_mismatch += mismatch_items
        total_pending += pending_items
    
    context = {
        'assignments': assignments,
        'total_verified': total_verified,
        'total_matched': total_matched,
        'total_mismatch': total_mismatch,
        'total_pending': total_pending,
        'today': timezone.now().date(),
    }
    return render(request, 'manager/verification_status.html', context)

@login_required
def manager_verification_details(request, assignment_id):
    """Manager view to see detailed verification status for an assignment"""
    if request.user.user_type != 'manager':
        return redirect('employee_dashboard')
    
    assignment = get_object_or_404(
        HardwareAssignment,
        id=assignment_id,
        assigned_by=request.user,
        actual_return_date__isnull=True
    )
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment)
    
    # Get serial entries for each item
    verification_details = []
    all_verified = True
    
    for item in items:
        try:
            serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
            # FIXED: Compare serial numbers for is_match
            is_match = serial_entry.serial_number == item.hardware.serial_number
            verification_details.append({
                'item': item,
                'serial_entry': serial_entry,
                'is_match': is_match,  # CORRECT: based on actual serial comparison
                'entered_serial': serial_entry.serial_number,
                'expected_serial': item.hardware.serial_number,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'entered_by': serial_entry.entered_by,
                'entered_at': serial_entry.entered_at,
                'verified_at': serial_entry.verified_at,
                'verified_by': serial_entry.verified_by
            })
            if not is_match:
                all_verified = False
        except HardwareSerialEntry.DoesNotExist:
            verification_details.append({
                'item': item,
                'serial_entry': None,
                'is_match': False,
                'entered_serial': None,
                'expected_serial': item.hardware.serial_number,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'entered_by': None,
                'entered_at': None,
                'verified_at': None,
                'verified_by': None
            })
            all_verified = False
    
    # Calculate statistics
    total_items = len(verification_details)
    verified_count = sum(1 for d in verification_details if d['verified_at'] is not None)  # Check verified_at instead of is_match
    mismatch_count = sum(1 for d in verification_details if d['entered_serial'] and d['entered_serial'] != d['expected_serial'])
    pending_count = sum(1 for d in verification_details if not d['entered_serial'])
    
    context = {
        'assignment': assignment,
        'verification_details': verification_details,
        'all_verified': all_verified,
        'total_items': total_items,
        'verified_count': verified_count,
        'mismatch_count': mismatch_count,
        'pending_count': pending_count,
        'verified_percentage': (verified_count / total_items * 100) if total_items > 0 else 0,
    }
    return render(request, 'manager/verification_details.html', context)
# ============== EMPLOYEE VIEWS ==============
@login_required
def employee_dashboard(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    current_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    # Calculate statistics
    current_assignments_count = current_assignments.count()
    completed_assignments_count = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=False
    ).count()
    
    # Get current exam city (from most recent assignment)
    current_exam_city = None
    if current_assignments.exists():
        current_exam_city = current_assignments.first().exam_city
    
    # Count pending serial entries
    pending_serials_count = 0
    for assignment in current_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        for item in items:
            if not hasattr(item, 'serial_entry'):
                pending_serials_count += 1
    
    # Get employee's full name
    employee_name = request.user.get_full_name() or request.user.username
    
    context = {
        'current_assignments': current_assignments,
        'current_assignments_count': current_assignments_count,
        'completed_assignments_count': completed_assignments_count,
        'pending_serials_count': pending_serials_count,
        'current_exam_city': current_exam_city,
        'employee_name': employee_name,  # Pass name to template
    }
    return render(request, 'employee/dashboard.html', context)
@login_required
def view_my_assignments(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignments = HardwareAssignment.objects.filter(employee=request.user).order_by('-assigned_date')
    
    # Add serial entry status to each assignment
    for assignment in assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        assignment.total_items = items.count()
        assignment.pending_serial_count = 0  # Count of items without serial entry
        assignment.entered_serial_count = 0  # Count of items with serial entry
        assignment.verified_serial_count = 0  # Count of verified serials
        
        for item in items:
            if hasattr(item, 'serial_entry'):
                assignment.entered_serial_count += 1
                if item.serial_entry.verified:
                    assignment.verified_serial_count += 1
            else:
                assignment.pending_serial_count += 1
    
    context = {
        'assignments': assignments,
    }
    return render(request, 'employee/view_my_assignments.html', context)

@login_required
def my_assignment_details(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    # Add serial entry info to each item
    for item in items:
        item.has_serial_entry = hasattr(item, 'serial_entry')
        if item.has_serial_entry:
            item.serial_number = item.serial_entry.serial_number
            item.is_verified = item.serial_entry.verified
        else:
            item.serial_number = None
            item.is_verified = False
    
    # Calculate statistics
    total_items = items.count()
    pending_count = sum(1 for item in items if not hasattr(item, 'serial_entry'))
    entered_count = total_items - pending_count
    verified_count = sum(1 for item in items if hasattr(item, 'serial_entry') and item.serial_entry.verified)
    
    context = {
        'assignment': assignment,
        'items': items,
        'total_items': total_items,
        'pending_count': pending_count,
        'entered_count': entered_count,
        'verified_count': verified_count,
    }
    return render(request, 'employee/my_assignment_details.html', context)
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from django.contrib.auth.decorators import login_required
from django.shortcuts import get_object_or_404, redirect
from .models import HardwareAssignment, HardwareAssignmentItem, HardwareType

@login_required
def export_assignment_excel(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    # Get all hardware types in the specific order you want
    hardware_type_order = [
        'Laptop', 'Firewall', 'PXE BOX', 'L1 Devices', 'Tatwiks', 
        'Cameras', 'Barcode Scanners', 'Chargers Laptop', 'Firwall Charger', 'PXE BOX Charger'
    ]
    
    # Get hardware types that exist in the database
    existing_hardware_types = HardwareType.objects.filter(name__in=hardware_type_order)
    
    # Create ordered list of hardware types
    hardware_types = []
    for type_name in hardware_type_order:
        try:
            hw_type = existing_hardware_types.get(name=type_name)
            hardware_types.append(hw_type)
        except HardwareType.DoesNotExist:
            # Create a dummy hardware type object if it doesn't exist
            from types import SimpleNamespace
            dummy_type = SimpleNamespace()
            dummy_type.name = type_name
            dummy_type.id = None
            hardware_types.append(dummy_type)
    
    # Group items by hardware type and collect serial numbers
    items_by_type = {}
    max_items_per_type = 0
    
    for type_name in hardware_type_order:
        type_items = [item for item in items if item.hardware.hardware_type.name == type_name]
        items_by_type[type_name] = type_items
        max_items_per_type = max(max_items_per_type, len(type_items))
    
    # Create a workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Convert UUID to string for sheet title (max 31 characters for Excel sheet names)
    assignment_id_str = str(assignment.assignment_id).replace('-', '')[:8]
    ws.title = f"Assign_{assignment_id_str}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="E04D00", end_color="E04D00", fill_type="solid")
    subheader_font = Font(bold=True, color="FFFFFF", size=11)
    subheader_fill = PatternFill(start_color="666666", end_color="666666", fill_type="solid")
    center_alignment = Alignment(horizontal="center", vertical="center")
    left_alignment = Alignment(horizontal="left", vertical="center")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Title Row - Assignment Information
    ws.merge_cells(f'A1:{get_column_letter(len(hardware_types))}1')
    title_cell = ws['A1']
    title_cell.value = f"ASSIGNMENT DETAILS - {assignment.assignment_id}"
    title_cell.font = Font(bold=True, size=14, color="E04D00")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30
    
    # Assignment Info Rows (compact)
    info_start_row = 3
    info_data = [
        ['Project:', assignment.project.project_name],
        ['Exam City:', assignment.exam_city or 'Not specified'],
        ['Employee:', assignment.employee.get_full_name() or assignment.employee.username],
        ['Assigned Date:', assignment.assigned_date.strftime('%d %b %Y')],
    ]
    
    for i, (label, value) in enumerate(info_data, start=info_start_row):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
        ws.row_dimensions[i].height = 18
    
    # Add spacing row
    spacing_row = info_start_row + len(info_data) + 1
    ws.row_dimensions[spacing_row].height = 10
    
    # Hardware Type Headers
    header_row = spacing_row + 1
    
    for col_idx, hw_type in enumerate(hardware_types, start=1):
        # Main header with hardware type name
        cell = ws.cell(row=header_row, column=col_idx, value=hw_type.name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_alignment
        cell.border = border
        
        # Subheader for serial numbers
        subheader_cell = ws.cell(row=header_row + 1, column=col_idx, value="Serial Numbers")
        subheader_cell.font = subheader_font
        subheader_cell.fill = subheader_fill
        subheader_cell.alignment = center_alignment
        subheader_cell.border = border
    
    # Add serial numbers under each hardware type
    data_start_row = header_row + 2
    
    for row_offset in range(max_items_per_type):
        current_row = data_start_row + row_offset
        
        for col_idx, hw_type in enumerate(hardware_types, start=1):
            type_items = items_by_type.get(hw_type.name, [])
            
            if row_offset < len(type_items):
                # This hardware type has an item at this position
                item = type_items[row_offset]
                
                # Get the serial number (entered or expected)
                if hasattr(item, 'serial_entry') and item.serial_entry.serial_number:
                    serial_value = item.serial_entry.serial_number
                else:
                    serial_value = item.hardware.serial_number
                
                cell = ws.cell(row=current_row, column=col_idx, value=serial_value)
                cell.border = border
                cell.alignment = center_alignment
                
                # Color code based on verification status
                if hasattr(item, 'serial_entry'):
                    if item.serial_entry.verified:
                        cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green for verified
                        cell.font = Font(bold=True, color="006100")
                    else:
                        cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow for pending
                        cell.font = Font(bold=True, color="9C5700")
                else:
                    cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red for not entered
                    cell.font = Font(bold=True, color="9C0006")
            else:
                # No item at this position - leave blank or put placeholder
                cell = ws.cell(row=current_row, column=col_idx, value="—")
                cell.border = border
                cell.alignment = center_alignment
                cell.font = Font(color="999999", italic=True)
    
    # Add summary section
    summary_row = data_start_row + max_items_per_type + 2
    ws.merge_cells(f'A{summary_row}:B{summary_row}')
    summary_title = ws.cell(row=summary_row, column=1, value="SUMMARY")
    summary_title.font = Font(bold=True, size=12)
    
    # Statistics
    total_items = items.count()
    verified_count = sum(1 for item in items if hasattr(item, 'serial_entry') and item.serial_entry.verified)
    pending_count = sum(1 for item in items if hasattr(item, 'serial_entry') and not item.serial_entry.verified)
    not_entered_count = total_items - (verified_count + pending_count)
    
    stats = [
        ['Total Items:', total_items],
        ['Verified:', verified_count],
        ['Pending:', pending_count],
        ['Not Entered:', not_entered_count],
    ]
    
    for i, (label, value) in enumerate(stats, start=summary_row + 1):
        ws[f'A{i}'] = label
        ws[f'B{i}'] = value
        ws[f'A{i}'].font = Font(bold=True)
    
    # Hardware type counts
    type_count_row = summary_row + len(stats) + 2
    ws.merge_cells(f'A{type_count_row}:B{type_count_row}')
    type_count_title = ws.cell(row=type_count_row, column=1, value="ITEMS PER TYPE")
    type_count_title.font = Font(bold=True, size=12)
    
    for i, hw_type in enumerate(hardware_types, start=type_count_row + 1):
        type_count = len(items_by_type.get(hw_type.name, []))
        ws[f'A{i}'] = hw_type.name
        ws[f'B{i}'] = type_count
        ws[f'A{i}'].font = Font(bold=True)
    
    # Auto-adjust column widths
    for col in range(1, len(hardware_types) + 1):
        column_letter = get_column_letter(col)
        max_length = len(hardware_types[col-1].name) + 5
        for row in range(data_start_row, data_start_row + max_items_per_type):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)) + 2)
        ws.column_dimensions[column_letter].width = min(max_length, 20)
    
    # Create the HttpResponse
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Convert UUID to string for filename
    assignment_id_str = str(assignment.assignment_id).replace('-', '')[:8]
    filename = f"assignment_{assignment_id_str}_{assignment.employee.username}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save the workbook to the response
    wb.save(response)
    return response
# ============== EMPLOYEE SERIAL NUMBER ENTRY ==============

@login_required
def enter_serial_numbers(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    
    # Check if assignment is already returned
    if assignment.actual_return_date:
        messages.error(request, 'This assignment has already been returned!')
        return redirect('employee_dashboard')
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    if request.method == 'POST':
        success_count = 0
        error_count = 0
        
        for item in items:
            serial_number = request.POST.get(f'serial_{item.id}')
            if serial_number and serial_number.strip():
                # Check if serial entry already exists
                if hasattr(item, 'serial_entry'):
                    # Update existing entry
                    serial_entry = item.serial_entry
                    serial_entry.serial_number = serial_number.strip()
                    serial_entry.save()
                    success_count += 1
                else:
                    # Create new entry
                    HardwareSerialEntry.objects.create(
                        assignment_item=item,
                        serial_number=serial_number.strip(),
                        entered_by=request.user
                    )
                    success_count += 1
            else:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'Successfully submitted {success_count} serial number(s)!')
        if error_count > 0:
            messages.warning(request, f'{error_count} item(s) were not submitted (empty serial number)')
        
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    # Pre-fill existing serial numbers
    for item in items:
        if hasattr(item, 'serial_entry'):
            item.existing_serial = item.serial_entry.serial_number
        else:
            item.existing_serial = ''
    
    context = {
        'assignment': assignment,
        'items': items,
    }
    return render(request, 'employee/enter_serial_numbers.html', context)

@login_required
def edit_serial_numbers(request, assignment_id):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    assignment = get_object_or_404(HardwareAssignment, id=assignment_id, employee=request.user)
    
    # Check if assignment is already returned
    if assignment.actual_return_date:
        messages.error(request, 'This assignment has already been returned!')
        return redirect('employee_dashboard')
    
    items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
    
    # Check if any serial has been verified (cannot edit verified serials)
    has_verified_serials = False
    for item in items:
        if hasattr(item, 'serial_entry') and item.serial_entry.verified:
            has_verified_serials = True
            break
    
    if has_verified_serials:
        messages.error(request, 'Cannot edit serial numbers that have been verified by manager!')
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    # Same as enter_serial_numbers view
    if request.method == 'POST':
        success_count = 0
        error_count = 0
        
        for item in items:
            serial_number = request.POST.get(f'serial_{item.id}')
            if serial_number and serial_number.strip():
                # Check if serial entry already exists
                if hasattr(item, 'serial_entry'):
                    # Update existing entry
                    serial_entry = item.serial_entry
                    serial_entry.serial_number = serial_number.strip()
                    serial_entry.save()
                    success_count += 1
                else:
                    # Create new entry
                    HardwareSerialEntry.objects.create(
                        assignment_item=item,
                        serial_number=serial_number.strip(),
                        entered_by=request.user
                    )
                    success_count += 1
            else:
                error_count += 1
        
        if success_count > 0:
            messages.success(request, f'Successfully updated {success_count} serial number(s)!')
        if error_count > 0:
            messages.warning(request, f'{error_count} item(s) were not updated (empty serial number)')
        
        return redirect('my_assignment_details', assignment_id=assignment_id)
    
    # Pre-fill existing serial numbers
    for item in items:
        if hasattr(item, 'serial_entry'):
            item.existing_serial = item.serial_entry.serial_number
        else:
            item.existing_serial = ''
    
    context = {
        'assignment': assignment,
        'items': items,
        'editing': True,
    }
    return render(request, 'employee/enter_serial_numbers.html', context)

# ============== API VIEWS ==============

@csrf_exempt
@login_required
def api_get_hardware_by_type(request):
    if request.method == 'GET':
        hardware_type_id = request.GET.get('type_id')
        
        if request.user.user_type == 'manager':
            hardware_items = Hardware.objects.filter(
                hardware_type_id=hardware_type_id,
                status='available',
                created_by=request.user
            ).values('id', 'serial_number', 'model_name', 'brand')
        else:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        return JsonResponse(list(hardware_items), safe=False)

@csrf_exempt
@login_required
def api_get_assignment_details(request, assignment_id):
    if request.method == 'GET':
        assignment = get_object_or_404(HardwareAssignment, id=assignment_id)
        
        if request.user.user_type == 'employee' and assignment.employee != request.user:
            return JsonResponse({'error': 'Unauthorized'}, status=403)
        
        items = HardwareAssignmentItem.objects.filter(assignment=assignment).select_related('hardware__hardware_type')
        
        data = {
            'assignment_id': str(assignment.assignment_id),
            'project': {
                'id': assignment.project.id,
                'name': assignment.project.project_name,
                'location': assignment.project.location,
            },
            'employee': assignment.employee.username,
            'assigned_by': assignment.assigned_by.username,
            'assigned_date': assignment.assigned_date.strftime('%Y-%m-%d'),
            'expected_return_date': assignment.expected_return_date.strftime('%Y-%m-%d'),
            'actual_return_date': assignment.actual_return_date.strftime('%Y-%m-%d') if assignment.actual_return_date else None,
            'notes': assignment.notes,
            'hardware_items': [
                {
                    'id': item.id,
                    'hardware_id': item.hardware.id,
                    'type': item.hardware.hardware_type.name,
                    'model': item.hardware.model_name,
                    'serial_number': item.hardware.serial_number,
                    'brand': item.hardware.brand,
                    'has_serial_entry': hasattr(item, 'serial_entry'),
                    'entered_serial': item.serial_entry.serial_number if hasattr(item, 'serial_entry') else None,
                    'verified': item.serial_entry.verified if hasattr(item, 'serial_entry') else False,
                }
                for item in items
            ]
        }
        
        return JsonResponse(data)

@csrf_exempt
@login_required
def api_check_serial_exists(request):
    if request.method == 'GET':
        serial_number = request.GET.get('serial_number')
        
        if not serial_number:
            return JsonResponse({'exists': False})
        
        exists = HardwareSerialEntry.objects.filter(serial_number=serial_number).exists()
        return JsonResponse({'exists': exists})

# ============== UTILITY VIEWS ==============

@login_required
def profile(request):
    context = {'user': request.user}
    return render(request, 'profile.html', context)

@login_required
def update_profile(request):
    if request.method == 'POST':
        user = request.user
        user.email = request.POST.get('email')
        user.phone = request.POST.get('phone')
        user.first_name = request.POST.get('first_name')
        user.last_name = request.POST.get('last_name')
        user.save()
        
        messages.success(request, 'Profile updated successfully!')
        return redirect('profile')
    
    context = {'user': request.user}
    return render(request, 'update_profile.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth import get_user_model
from django.utils import timezone
from django.core.mail import send_mail
from django.conf import settings
from django.template.loader import render_to_string
from django.contrib.auth.hashers import make_password
from .models import PasswordResetOTP
import uuid

User = get_user_model()

def forgot_password(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        
        try:
            user = User.objects.get(username=username)
            
            # Generate OTP
            otp_obj = PasswordResetOTP.generate_otp(user)
            
            # Send email with OTP
            subject = 'Password Reset OTP - Eduquity Hardware Management'
            html_message = render_to_string('auth/password_reset_email.html', {
                'user': user,
                'otp': otp_obj.otp,
                'expiry_minutes': 5,
                'token': otp_obj.token,
            })
            
            send_mail(
                subject=subject,
                message=f'Your OTP for password reset is: {otp_obj.otp}. This OTP is valid for 5 minutes.',
                from_email=settings.DEFAULT_FROM_EMAIL,
                recipient_list=[user.email],
                html_message=html_message,
                fail_silently=False,
            )
            
            messages.success(request, f'OTP has been sent to your registered email address ({user.email}). Please check your inbox.')
            return redirect('verify_otp', token=otp_obj.token)
            
        except User.DoesNotExist:
            messages.error(request, 'No account found with this Employee ID.')
        except Exception as e:
            messages.error(request, f'Failed to send OTP. Error: {str(e)}')
    
    return render(request, 'auth/forgot_password.html')

def verify_otp(request, token):
    try:
        otp_obj = PasswordResetOTP.objects.get(token=token, is_used=False)
        
        if otp_obj.is_expired():
            messages.error(request, 'OTP has expired. Please request a new one.')
            return redirect('forgot_password')
        
        if request.method == 'POST':
            entered_otp = request.POST.get('otp')
            
            if entered_otp == otp_obj.otp:
                otp_obj.is_used = True
                otp_obj.save()
                return redirect('reset_password', token=token)
            else:
                messages.error(request, 'Invalid OTP. Please try again.')
        
        return render(request, 'auth/verify_otp.html', {
            'token': token,
            'email': otp_obj.user.email[:3] + '*****' + otp_obj.user.email[otp_obj.user.email.find('@'):]
        })
    
    except PasswordResetOTP.DoesNotExist:
        messages.error(request, 'Invalid or expired OTP link.')
        return redirect('forgot_password')
def reset_password(request, token):
    try:
        otp_obj = PasswordResetOTP.objects.get(token=token)
        
        if otp_obj.is_used == False:
            messages.error(request, 'Please verify OTP first.')
            return redirect('verify_otp', token=token)
        
        if request.method == 'POST':
            password = request.POST.get('password')
            confirm_password = request.POST.get('confirm_password')
            
            if password != confirm_password:
                messages.error(request, 'Passwords do not match!')
                return render(request, 'auth/reset_password.html', {'token': token})
            
            if len(password) < 8:
                messages.error(request, 'Password must be at least 8 characters long.')
                return render(request, 'auth/reset_password.html', {'token': token})
            
            # Update user password
            user = otp_obj.user
            user.set_password(password)
            
            # If employee is resetting password first time
            if user.user_type == 'employee' and user.is_first_login:
                user.is_first_login = False
            
            user.save()
            
            # Send confirmation email (with error handling)
            try:
                subject = 'Password Reset Successful - Eduquity Hardware Management'
                
                # Check if template exists before trying to render
                try:
                    html_message = render_to_string('auth/password_reset_success_email.html', {
                        'user': user,
                        'now': timezone.now(),
                    })
                except TemplateDoesNotExist:
                    # Create a simple HTML message if template doesn't exist
                    html_message = f'''
                    <!DOCTYPE html>
                    <html>
                    <body style="font-family: Arial, sans-serif; line-height: 1.6;">
                        <h2>Password Reset Successful</h2>
                        <p>Hello {user.username},</p>
                        <p>Your password has been successfully reset for the Eduquity Hardware Management System.</p>
                        <p>If you did not request this password reset, please contact your system administrator immediately.</p>
                        <p><strong>Eduquity Hardware Management Team</strong></p>
                    </body>
                    </html>
                    '''
                
                send_mail(
                    subject=subject,
                    message=f'Your password has been reset successfully on {timezone.now().strftime("%B %d, %Y at %I:%M %p")}.',
                    from_email=settings.DEFAULT_FROM_EMAIL,
                    recipient_list=[user.email],
                    html_message=html_message,
                    fail_silently=False,
                )
                
            except Exception as e:
                # Log the error but continue
                print(f"Failed to send password reset email: {str(e)}")
                # Don't show error to user since password was reset successfully
            
            messages.success(request, 'Password reset successfully! You can now login with your new password.')
            return redirect('login')
        
        return render(request, 'auth/reset_password.html', {'token': token})
    
    except PasswordResetOTP.DoesNotExist:
        messages.error(request, 'Invalid reset link.')
        return redirect('forgot_password')
    
@login_required
def my_hardware(request):
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    # Get active assignments
    active_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    # Get completed assignments
    completed_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=False
    ).order_by('-assigned_date')[:5]  # Last 5 completed
    
    total_items = 0
    verified_count = 0
    pending_count = 0
    matched_count = 0
    mismatch_count = 0
    
    # Process each assignment
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        assignment.hardware_count = items.count()
        total_items += assignment.hardware_count
        
        assignment.verified_count = 0
        assignment.matched_count = 0
        assignment.mismatch_count = 0
        assignment.pending_count = 0
        # FIXED: Use a different name instead of hardware_items
        assignment.items_list = []  # Changed from hardware_items to items_list
        
        for item in items:
            hardware_data = {
                'id': item.id,
                'hardware_type': item.hardware.hardware_type.name,
                'model': item.hardware.model_name,
                'brand': item.hardware.brand,
                'expected_serial': item.hardware.serial_number,
                'status': item.hardware.status,
            }
            
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                hardware_data['serial_number'] = serial_entry.serial_number
                hardware_data['verified'] = serial_entry.verified
                hardware_data['verified_by'] = serial_entry.verified_by
                hardware_data['verified_at'] = serial_entry.verified_at
                hardware_data['entered_at'] = serial_entry.entered_at
                
                if serial_entry.verified:
                    assignment.verified_count += 1
                    verified_count += 1
                else:
                    if serial_entry.serial_number == item.hardware.serial_number:
                        assignment.matched_count += 1
                        matched_count += 1
                    else:
                        assignment.mismatch_count += 1
                        mismatch_count += 1
                        
            except HardwareSerialEntry.DoesNotExist:
                hardware_data['serial_number'] = None
                hardware_data['verified'] = False
                assignment.pending_count += 1
                pending_count += 1
            
            # FIXED: Use the new variable name
            assignment.items_list.append(hardware_data)
    
    # Process completed assignments
    for assignment in completed_assignments:
        assignment.hardware_count = HardwareAssignmentItem.objects.filter(assignment=assignment).count()
    
    context = {
        'assignments': active_assignments,
        'completed_assignments': completed_assignments,
        'total_items': total_items,
        'verified_count': verified_count,
        'pending_count': pending_count,
        'matched_count': matched_count,
        'mismatch_count': mismatch_count,
        'active_assignments': active_assignments.count(),
    }
    return render(request, 'employee/my_hardware.html', context)

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

@login_required
def export_my_hardware_excel(request):
    """Export employee's hardware data to Excel with filename: examcity_employeename_date.xlsx"""
    if request.user.user_type != 'employee':
        return redirect('manager_dashboard')
    
    # Get active assignments
    active_assignments = HardwareAssignment.objects.filter(
        employee=request.user,
        actual_return_date__isnull=True
    ).order_by('-assigned_date')
    
    # Get employee's current exam city (from most recent assignment)
    exam_city = "NoCity"
    if active_assignments.exists():
        exam_city = active_assignments.first().exam_city.replace(" ", "_")
    
    # Get employee name
    employee_name = request.user.get_full_name() or request.user.username
    employee_name = employee_name.replace(" ", "_")
    
    # Get current date
    current_date = datetime.now().strftime("%Y%m%d")
    
    # Create filename in format: examcity_employeename_date.xlsx
    filename = f"{exam_city}_{employee_name}_{current_date}.xlsx"
    
    # Create Excel workbook and sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "My Hardware Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    success_fill = PatternFill(start_color="DFF0D8", end_color="DFF0D8", fill_type="solid")  # Light green
    info_fill = PatternFill(start_color="D9EDF7", end_color="D9EDF7", fill_type="solid")      # Light blue
    warning_fill = PatternFill(start_color="FCF8E3", end_color="FCF8E3", fill_type="solid")    # Light yellow
    danger_fill = PatternFill(start_color="F2DEDE", end_color="F2DEDE", fill_type="solid")     # Light red
    
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # Write header row
    headers = [
        'Assignment ID', 'Project', 'Exam City', 'Assigned Date', 'Expected Return',
        'Hardware Type', 'Model', 'Brand', 'Assigned Serial', 'Entered Serial',
        'Entry Status', 'Verification Status', 'Verified By', 'Verified On'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write data rows
    row_num = 2
    total_items = 0
    
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        
        for item in items:
            total_items += 1
            
            # Get serial entry if exists
            serial_entry = None
            entered_serial = "Not entered"
            entry_status = "Pending"
            verification_status = "Pending"
            verified_by = "-"
            verified_on = "-"
            
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                entered_serial = serial_entry.serial_number
                
                if serial_entry.verified:
                    verification_status = "Verified"
                    verified_by = serial_entry.verified_by.get_full_name() or serial_entry.verified_by.username if serial_entry.verified_by else "-"
                    verified_on = serial_entry.verified_at.strftime("%d/%m/%Y %H:%M") if serial_entry.verified_at else "-"
                    
                    if serial_entry.serial_number == item.hardware.serial_number:
                        entry_status = "Verified - Correct"
                    else:
                        entry_status = "Verified - Mismatch"
                else:
                    if serial_entry.serial_number == item.hardware.serial_number:
                        entry_status = "Matched - Pending"
                        verification_status = "Pending Verification"
                    else:
                        entry_status = "Mismatch"
                        verification_status = "Not Verified"
                        
            except HardwareSerialEntry.DoesNotExist:
                entry_status = "Not Entered"
                verification_status = "Pending"
            
            # Determine row color based on status
            if serial_entry and serial_entry.verified:
                row_fill = success_fill
            elif serial_entry and serial_entry.serial_number == item.hardware.serial_number:
                row_fill = info_fill
            elif serial_entry and serial_entry.serial_number != item.hardware.serial_number:
                row_fill = danger_fill
            else:
                row_fill = warning_fill
            
            row_data = [
                str(assignment.assignment_id)[:8],
                assignment.project.project_name,
                assignment.exam_city,
                assignment.assigned_date.strftime("%d/%m/%Y"),
                assignment.expected_return_date.strftime("%d/%m/%Y"),
                item.hardware.hardware_type.name,
                item.hardware.model_name,
                item.hardware.brand or "-",
                item.hardware.serial_number,
                entered_serial,
                entry_status,
                verification_status,
                verified_by,
                verified_on
            ]
            
            for col_num, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col_num, value=value)
                cell.border = border
                cell.fill = row_fill
                cell.alignment = Alignment(vertical="center")
                
                # Format date columns
                if col_num in [4, 5, 14]:  # Date columns
                    cell.alignment = Alignment(horizontal="center", vertical="center")
            
            row_num += 1
    
    # Add summary section
    row_num += 2
    summary_row = row_num
    
    # Calculate totals
    verified_count = 0
    matched_count = 0
    mismatch_count = 0
    pending_count = 0
    
    for assignment in active_assignments:
        items = HardwareAssignmentItem.objects.filter(assignment=assignment)
        for item in items:
            try:
                serial_entry = HardwareSerialEntry.objects.get(assignment_item=item)
                if serial_entry.verified:
                    verified_count += 1
                elif serial_entry.serial_number == item.hardware.serial_number:
                    matched_count += 1
                else:
                    mismatch_count += 1
            except HardwareSerialEntry.DoesNotExist:
                pending_count += 1
    
    # Write summary
    summary_headers = ['Summary Statistics', 'Value']
    summary_data = [
        ['Total Items', total_items],
        ['Verified Items', verified_count],
        ['Matched Items (Pending)', matched_count],
        ['Mismatched Items', mismatch_count],
        ['Pending Entry', pending_count],
        ['Generated On', datetime.now().strftime("%d/%m/%Y %H:%M:%S")],
        ['Employee Name', request.user.get_full_name() or request.user.username],
        ['Employee Email', request.user.email],
        ['Exam City', assignment.exam_city if active_assignments.exists() else 'Not Assigned'],
    ]
    
    # Style summary header
    for col_num, header in enumerate(summary_headers, 1):
        cell = ws.cell(row=summary_row, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Write summary data
    for i, (label, value) in enumerate(summary_data, summary_row + 1):
        label_cell = ws.cell(row=i, column=1, value=label)
        label_cell.border = border
        label_cell.font = Font(bold=True)
        
        value_cell = ws.cell(row=i, column=2, value=value)
        value_cell.border = border
        value_cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in range(1, row_num):
            cell_value = ws.cell(row=row, column=col).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        
        adjusted_width = min(max_length + 4, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response with the new filename format
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response

